import openpyxl
import pandas as pd
path=r'D:\Python\Mapping Doc.xlsx'


df=pd.read_excel(r'D:\Python\NewMappingDoc.xlsx',sheet_name="MySheet")
print(df)

df.to_excel(r'D:\Python\NewMappingDoc2.xlsx',sheet_name="Sheet1",index=False)

# To open workbook, workbook object is created
wb=openpyxl.load_workbook(path)
sheet_names=wb.sheetnames
active_sheet_object=wb.active
print(sheet_names)

print(active_sheet_object.cell(10,1).value)

print("Active_Sheet=",wb.active.title)

d=dict()
for name in sheet_names:
    ws=wb[name]
    print(ws)
    rows=ws.max_row
    cols=ws.max_column
    d[name]={"rows":rows,"cols":cols}
    print("***************************")

    # for i in range(1,rows+1):
    #     for j in range(1,cols+1):
    #         pass
    #         print(ws.cell(row=i,column=j).value,end=" , ")
    #     print()

    for row in ws.iter_rows(min_row=1, max_row=5, min_col=1, max_col=3):
        for cell in row:
            # print(cell.column_letter,end=":")
            print(cell.value, end=" , ")
        print()

    # display row data in tuble format
    # for row in ws.iter_rows(values_only=True):
    #     print(row)

 
print(d)

column_A=wb.active['B']
set=set()
for cell in column_A:
    set.add(cell.value)
print(set)

# simple step to read a column data
column_data=[wb.active.cell(row=i,column=3).value for i in range(2,13)]
print(column_data)

row_data=[wb.active.cell(row=5,column=i).value for i in range(1,5)]
print(row_data)

# row_column_data={wb.active.cell(row=i,column=j).value for i,j in range(1,5)}
# print(row_column_data)
wb.close()

# Create excel file and add data
new_wb=openpyxl.Workbook()
new_ws=new_wb.active
new_ws.title="MySheet"
new_ws['A1']="ID"
new_ws['B1']="Name" 
new_ws['C1']="City"
new_ws.append([1,"Alice","New York"])
new_ws.append([2,"Bob","Los Angeles"])
new_ws.append([3,"Charlie","Chicago"])
row=(4,"David","Houston")
new_ws.append(row)
new_wb.save(r'D:\Python\NewMappingDoc.xlsx')
new_wb.close()

